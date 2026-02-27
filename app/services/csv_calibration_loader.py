import os
import re
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class CSVCalibrationLoader:
    
    FEATURE_COLUMN_CANDIDATES = ["name", "module name", "feature", "module"]
    
    TOTAL_HOURS_CANDIDATES = ["total hours", "total", "hours"]
    
    COMPONENT_HOUR_COLUMNS = [
        "web mobile",
        "backend",
        "wireframe",
        "visual design"
    ]
    
    SKIP_ROW_KEYWORDS = [
        "total", "subtotal", "summary", "grand total", 
        "sub-total", "sub total", "grand-total"
    ]
    
    def __init__(self, calibration_folder: str = "app/data/calibration"):
        self.calibration_folder = calibration_folder
        self.calibration_data: Dict[str, Dict] = {}
    
    def load_all_calibrations(self) -> Dict[str, Dict]:
        """
        Load all Excel files from calibration folder and aggregate data.
        
        Returns:
            Dict mapping normalized feature names to calibration data
        """
        if not os.path.exists(self.calibration_folder):
            logger.warning(f"Calibration folder not found: {self.calibration_folder}")
            return {}
        
        excel_files = list(Path(self.calibration_folder).glob("*.xlsx"))
        excel_files.extend(Path(self.calibration_folder).glob("*.xls"))
        
        if not excel_files:
            logger.info("No Excel files found in calibration folder")
            return {}
        
        logger.info(f"Found {len(excel_files)} Excel file(s) to process")
        
        all_records: List[Tuple[str, float, str]] = []
        
        for excel_file in excel_files:
            try:
                records = self._process_excel_file(excel_file)
                all_records.extend(records)
                logger.info(f"Processed {excel_file.name}: {len(records)} records")
            except Exception as e:
                logger.error(f"Failed to process {excel_file.name}: {str(e)}")
                continue
        
        self.calibration_data = self._aggregate_records(all_records)
        
        logger.info(f"Loaded {len(self.calibration_data)} unique features from calibration data")
        
        return self.calibration_data
    
    def _process_excel_file(self, file_path: Path) -> List[Tuple[str, float, str]]:
        """
        Process all sheets in an Excel file.
        Tries openpyxl first; falls back to calamine for files with invalid/corrupt stylesheet XML.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of tuples (feature_name, hours, source)
        """
        records = []
        sheet_names: List[str] = []
        engine: Optional[str] = None

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
        except Exception as e:
            logger.warning(
                f"Openpyxl failed for {file_path.name} (e.g. invalid stylesheet XML): {e}. "
                "Trying calamine engine..."
            )
            try:
                excel_file = pd.ExcelFile(file_path, engine="calamine")
                sheet_names = excel_file.sheet_names
                engine = "calamine"
            except ImportError:
                logger.error(
                    f"python-calamine not installed. Install with: pip install python-calamine. "
                    f"Skipping {file_path.name}"
                )
                return []
            except Exception as calamine_err:
                logger.error(f"Failed to open workbook {file_path.name}: {calamine_err}")
                raise

        if not sheet_names:
            logger.warning(f"No sheets found in {file_path.name}")
            return []

        logger.info(f"Processing {file_path.name} with {len(sheet_names)} sheet(s)")

        for sheet_name in sheet_names:
            try:
                sheet_records = self._process_sheet(file_path, sheet_name, engine=engine)
                if sheet_records:
                    records.extend(sheet_records)
                    logger.debug(f"  ✓ {sheet_name}: {len(sheet_records)} records")
                else:
                    logger.debug(f"  ✗ {sheet_name}: skipped (no valid data)")
            except Exception as e:
                logger.warning(f"  ✗ {sheet_name}: error - {str(e)}")
                continue

        return records
    
    def _process_sheet(
        self,
        file_path: Path,
        sheet_name: str,
        engine: Optional[str] = None,
    ) -> List[Tuple[str, float, str]]:
        """
        Process a single sheet and extract feature-hour pairs.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to process
            engine: Optional engine to use ('openpyxl', 'calamine', 'xlrd'). If None, tries openpyxl then xlrd.
            
        Returns:
            List of tuples (feature_name, hours, source)
        """
        df = None
        if engine:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
            except Exception as e:
                logger.debug(f"Cannot read sheet {sheet_name} with {engine}: {str(e)}")
                return []
        else:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            except Exception:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine="xlrd")
                except Exception as e:
                    logger.debug(f"Cannot read sheet {sheet_name}: {str(e)}")
                    return []

        if df is None:
            return []
        
        if df.empty or len(df.columns) == 0:
            return []
        
        df.columns = [str(col).strip().lower() for col in df.columns]
        
        feature_col = self._find_feature_column(df)
        if not feature_col:
            return []
        
        hours_col, component_cols = self._find_hours_columns(df)
        
        if not hours_col and not component_cols:
            return []
        
        records = []
        source = f"{file_path.name}:{sheet_name}"
        
        for idx, row in df.iterrows():
            feature_name = self._extract_feature_name(row, feature_col)
            
            if not feature_name:
                continue
            
            if self._should_skip_row(feature_name):
                continue
            
            hours = self._extract_hours(row, hours_col, component_cols)
            
            if hours <= 0:
                continue
            
            records.append((feature_name, hours, source))
        
        return records
    
    def _find_feature_column(self, df: pd.DataFrame) -> Optional[str]:
        """
        Find the feature name column.
        
        Args:
            df: DataFrame with lowercase column names
            
        Returns:
            Column name or None
        """
        for col in df.columns:
            for candidate in self.FEATURE_COLUMN_CANDIDATES:
                if candidate in col:
                    return col
        return None
    
    def _find_hours_columns(self, df: pd.DataFrame) -> Tuple[Optional[str], List[str]]:
        """
        Find hours columns (total or components).
        
        Args:
            df: DataFrame with lowercase column names
            
        Returns:
            Tuple of (total_hours_col, component_cols_list)
        """
        total_col = None
        for col in df.columns:
            for candidate in self.TOTAL_HOURS_CANDIDATES:
                if candidate in col:
                    total_col = col
                    break
            if total_col:
                break
        
        component_cols = []
        for col in df.columns:
            for component in self.COMPONENT_HOUR_COLUMNS:
                if component in col:
                    component_cols.append(col)
                    break
        
        return total_col, component_cols
    
    def _extract_feature_name(self, row: pd.Series, feature_col: str) -> Optional[str]:
        """
        Extract and validate feature name from row.
        
        Args:
            row: DataFrame row
            feature_col: Feature column name
            
        Returns:
            Cleaned feature name or None
        """
        try:
            value = row[feature_col]
            if pd.isna(value):
                return None
            
            name = str(value).strip()
            
            if not name or len(name) < 2:
                return None
            
            return name
        except Exception:
            return None
    
    def _should_skip_row(self, feature_name: str) -> bool:
        """
        Check if row should be skipped based on feature name.
        
        Args:
            feature_name: Feature name to check
            
        Returns:
            True if row should be skipped
        """
        name_lower = feature_name.lower()
        
        for keyword in self.SKIP_ROW_KEYWORDS:
            if keyword in name_lower:
                return True
        
        return False
    
    def _extract_hours(
        self, 
        row: pd.Series, 
        total_col: Optional[str], 
        component_cols: List[str]
    ) -> float:
        """
        Extract hours from row using priority logic.
        
        Args:
            row: DataFrame row
            total_col: Total hours column name (if exists)
            component_cols: Component hour column names
            
        Returns:
            Extracted hours (0 if invalid)
        """
        if total_col:
            try:
                value = row[total_col]
                if pd.notna(value):
                    hours = float(value)
                    if hours > 0:
                        return hours
            except (ValueError, TypeError):
                pass
        
        if component_cols:
            total = 0.0
            for col in component_cols:
                try:
                    value = row[col]
                    if pd.notna(value):
                        total += float(value)
                except (ValueError, TypeError):
                    continue
            
            return total
        
        return 0.0
    
    def _normalize_feature_name(self, name: str) -> str:
        """
        Normalize feature name for consistent matching.
        Must match CalibrationEngine normalization.
        
        Args:
            name: Raw feature name
            
        Returns:
            Normalized feature name (lowercase, alphanumeric only)
        """
        normalized = re.sub(r'[^a-z0-9]+', '', name.lower())
        return normalized
    
    def _aggregate_records(
        self, 
        records: List[Tuple[str, float, str]]
    ) -> Dict[str, Dict]:
        """
        Aggregate records into weighted averages.
        
        Args:
            records: List of (feature_name, hours, source) tuples
            
        Returns:
            Dict mapping normalized feature to aggregated data
        """
        aggregated: Dict[str, Dict] = {}
        
        for feature_name, hours, source in records:
            normalized_name = self._normalize_feature_name(feature_name)
            
            if not normalized_name:
                continue
            
            if normalized_name not in aggregated:
                aggregated[normalized_name] = {
                    "total_hours": 0.0,
                    "sample_size": 0,
                    "sources": []
                }
            
            aggregated[normalized_name]["total_hours"] += hours
            aggregated[normalized_name]["sample_size"] += 1
            
            if source not in aggregated[normalized_name]["sources"]:
                aggregated[normalized_name]["sources"].append(source)
        
        for feature_data in aggregated.values():
            feature_data["avg_hours"] = round(
                feature_data["total_hours"] / feature_data["sample_size"], 
                1
            )
        
        return aggregated
    
    def get_calibration_data(self) -> Dict[str, Dict]:
        """
        Get loaded calibration data.
        
        Returns:
            Calibration data dictionary
        """
        return self.calibration_data
